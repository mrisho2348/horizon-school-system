# views.py
from decimal import Decimal, InvalidOperation
import logging
from django.http import HttpResponseRedirect
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.db import IntegrityError
from django.urls import reverse
import openpyxl
from openpyxl import load_workbook
from result_module.models import ClassLevel, ExamType, Result, Students, Subject, SujbectWiseResults
from .resources import ExamTypeResources, SubjectResources, SujbectWiseResultsResources
from .forms import ExcelUploadForm, ImportExamTypeForm, ImportStudentsForm, ImportSubjectForm, ImportSujbectWiseResultsForm
from tablib import Dataset
logger = logging.getLogger(__name__)



def import_student_records(request):
    if request.method == 'POST':
        form = ImportStudentsForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            try:
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active

                # Read headers
                headers = [cell.value.strip() for cell in sheet[1]]  # Trim headers
                required_headers = [
                    'first_name', 'middle_name', 'last_name', 
                    'physical_disabilities_condition', 'current_class', 
                    'date_of_birth', 'gender', 'first_phone_number', 'second_phone_number', 
                    'fee_payer_number', 'address', 'branch'
                ]

                # Validate headers
                if headers[:len(required_headers)] != required_headers:
                    messages.error(request, 'Invalid file format')
                    return render(request, 'hod_template/import_students.html', {'form': form})

                # Read data from rows
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    data = dict(zip(headers, (cell.strip() if isinstance(cell, str) else cell for cell in row)))  # Trim data
                    try:
                        Students.objects.create(                            
                            first_name=data['first_name'],
                            middle_name=data['middle_name'],
                            last_name=data['last_name'],
                            physical_disabilities_condition=data.get('physical_disabilities_condition', ''),
                            current_class=data.get('current_class', ''),
                            date_of_birth=data.get('date_of_birth'),
                            gender=data.get('gender', ''),
                            first_phone_number=data.get('first_phone_number', ''),
                            second_phone_number=data.get('second_phone_number', ''),
                            fee_payer_number=data.get('fee_payer_number', ''),
                            address=data.get('address', ''),
                            branch=data.get('branch', ''),                     
                        )
                    except IntegrityError:
                        # Skip duplicate entries and continue
                        continue

                return HttpResponseRedirect(reverse('manage_student'))

            except Exception as e:
                messages.error(request, f"Failed to import data: {str(e)}")
                return render(request, 'hod_template/import_students.html', {'form': form})
    else:
        form = ImportStudentsForm()

    return render(request, 'hod_template/import_students.html', {'form': form})



def import_student_results(request, exam_type_id, class_level_id):
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            exam_date = form.cleaned_data['date_of_exam']
            
            # Load the workbook and select the first sheet
            wb = load_workbook(filename=excel_file, data_only=True)
            ws = wb.active
            
            # Read the headers
            headers = [cell.value for cell in ws[1]]
            try:
                registration_column = headers.index('Student Registration Number')
            except ValueError:
                messages.error(request, "The Excel file must have a 'Student Registration Number' column.")
                return redirect('admin_import_student_results', exam_type_id=exam_type_id, class_level_id=class_level_id)

            # Get exam type and class level
            exam_type = get_object_or_404(ExamType, id=exam_type_id)
            class_level = get_object_or_404(ClassLevel, id=class_level_id)

            # Process each row of the Excel sheet
            for row in ws.iter_rows(min_row=2, values_only=True):
                registration_number = row[registration_column]
                if not registration_number:
                    continue
                
                student = Students.objects.filter(registration_number=registration_number).first()
                
                if student:
                    # Collect subject scores
                    for index, score in enumerate(row):
                        if index == registration_column or index == registration_column + 1:
                            continue  # Skip the student registration and name columns

                        subject_name = headers[index]
                        subject = Subject.objects.filter(subject_name=subject_name).first()

                        if subject:
                            # Validate score
                            if score is not None and (score < 0 or score > 100):
                                continue  # Skip invalid scores

                            # Create or update the result
                            Result.objects.update_or_create(
                                student=student,
                                subject=subject,
                                exam_type=exam_type,
                                class_level=class_level,
                                date_of_exam=exam_date,
                                defaults={'marks': score}
                            )
                            
            messages.success(request, 'Student results imported successfully!')
            return redirect('admin_view_student_to_add_result', exam_type_id=exam_type_id, class_level_id=class_level_id)
    else:
        form = ExcelUploadForm()
    
    context = {'form': form}
    return render(request, 'hod_template/import_student_results.html', context)
    

def import_subject_records(request):
    if request.method == 'POST':
        form = ImportSubjectForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                resource = SubjectResources()
                new_records = request.FILES['student_file']
                
                # Use tablib to load the imported data
                dataset = resource.export()
                imported_data = dataset.load(new_records.read(), format='xlsx')  # Assuming you are using xlsx, adjust accordingly

                for data in imported_data:
                     subject_recode = Subject(
                        subject_name=data[0],
                                      
                    )
                     subject_recode.save()

                return redirect('manage_subject') 
            except Exception as e:
                messages.error(request, f'An error occurred: {e}')

    else:
        form = ImportSubjectForm()

    return render(request, 'hod_template/import_subject.html', {'form': form})

def import_exam_type_records(request):
    if request.method == 'POST':
        form = ImportExamTypeForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                resource = ExamTypeResources()
                new_records = request.FILES['student_file']
                
                # Use tablib to load the imported data
                dataset = resource.export()
                imported_data = dataset.load(new_records.read(), format='xlsx')  # Assuming you are using xlsx, adjust accordingly

                for data in imported_data:
                     subject_recode = ExamType(
                        name=data[0],
                        description=data[0],
                                      
                    )
                     subject_recode.save()

                return redirect('manage_exam_type') 
            except Exception as e:
                messages.error(request, f'An error occurred: {e}')

    else:
        form = ImportExamTypeForm()

    return render(request, 'hod_template/import_exam_type.html', {'form': form})


