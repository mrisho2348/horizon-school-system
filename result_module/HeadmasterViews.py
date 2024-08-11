
from collections import defaultdict
from io import BytesIO
import json
from django import forms
from django.views.generic import CreateView, UpdateView
from django.urls import reverse_lazy, reverse
from django.db import IntegrityError
from django.http import Http404, HttpResponse,  HttpResponseRedirect, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView
from django.views.decorators.http import require_GET,require_POST
from result_module.forms import AddStaffForm, ClassFeeForm, ClassLevelForm, EditStaffForm, ExpenditureForm, FeePaymentForm, FeeStructureForm, MadrasatulFeeForm, MadrasatulFeePaymentForm, ResultForm, SchoolFeesInstallmentForm, StudentForm, StudentResultForm, SubjectForm, TransportFeeForm, TransportFeePaymentForm
from result_module.models import AdminHOD,  ClassAttendance, ClassFee, ClassLevel, CustomUser, ExamMetrics, ExamType, Expenditure, FeePayment, FeeStructure, FeedBackStaff,  LeaveReportStaffs, MadrasatulFee, MadrasatulFeePayment,  Result, SchoolFeesInstallment, SchoolFeesPayment,  Staffs, StudentClassAttendance, StudentExamInfo, StudentPositionInfo, Students, Subject, SujbectWiseResults, TransportFee, TransportFeePayment
from django.contrib import messages
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.template.loader import render_to_string
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, F
from django.db import models
from django.db.models import Q
from django.db.models import Count
from django.core.mail import send_mail
from django.core.exceptions import ValidationError
from django.core.files.storage import FileSystemStorage
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl import Workbook
# Create your views here.
@login_required
def headmaster_home(request):
    try:
        # Fetch the branch of the currently logged-in staff member
        staff_branch = request.user.staffs.branch  # Assuming `staffs` is the related name for the Staffs model

        # Fetch total number of students in the same branch as the logged-in staff
        total_students = Students.objects.filter(branch=staff_branch).count()
    except Exception as e:
        return render(request, "headmaster_template/error.html", {"error_message": f"Error fetching total students: {str(e)}"})  
    
    try:
        # Fetch total number of subjects (assuming subjects are not branch-specific)
        total_subjects = Subject.objects.count()
    except Exception as e:
        return render(request, "headmaster_template/error.html", {"error_message": f"Error fetching total subjects: {str(e)}"})

    try:
        # Fetch total number of female students in the same branch as the logged-in staff
        total_female_students = Students.objects.filter(gender='Female', branch=staff_branch).count()
    except Exception as e:
        return render(request, "headmaster_template/error.html", {"error_message": f"Error fetching total female students: {str(e)}"})

    try:
        # Fetch total number of male students in the same branch as the logged-in staff
        total_male_students = Students.objects.filter(gender='Male', branch=staff_branch).count()
    except Exception as e:
        return render(request, "headmaster_template/error.html", {"error_message": f"Error fetching total male students: {str(e)}"})

    try:
        # Counting the total number of students in the same branch as the logged-in staff
        students_count = Students.objects.filter(branch=staff_branch).count()
    except Exception as e:
        return render(request, "headmaster_template/error.html", {"error_message": f"Error counting total students: {str(e)}"})

    try:
        # Calculating total fees collected in the same branch as the logged-in staff
        total_school_fees = SchoolFeesPayment.objects.filter(student__branch=staff_branch).aggregate(total_collected=Sum('amount_paid'))['total_collected'] or 0
        total_madrasatul_fees = MadrasatulFeePayment.objects.filter(student__branch=staff_branch).aggregate(total_collected=Sum('amount_paid'))['total_collected'] or 0
        total_transport_fees = TransportFeePayment.objects.filter(student__branch=staff_branch).aggregate(total_collected=Sum('amount_paid'))['total_collected'] or 0
        total_fee_payment = FeePayment.objects.filter(student__branch=staff_branch).aggregate(total_collected=Sum('amount_paid'))['total_collected'] or 0

        total_fees_collected = (total_school_fees + total_madrasatul_fees +
                                total_transport_fees + total_fee_payment)
    except Exception as e:
        return render(request, "headmaster_template/error.html", {"error_message": f"Error calculating total fees collected: {str(e)}"})

    try:
        # Counting pending payments in the same branch as the logged-in staff
        pending_payments_count = SchoolFeesPayment.objects.filter(student__branch=staff_branch, amount_remaining__gt=0).count()
    except Exception as e:
        return render(request, "headmaster_template/error.html", {"error_message": f"Error counting pending payments: {str(e)}"})

    try:
        # Calculating total expenditure in the same branch as the logged-in staff
        total_expenditure = Expenditure.objects.filter(branch=staff_branch).aggregate(total_spent=Sum('amount'))['total_spent'] or 0
    except Exception as e:
        return render(request, "headmaster_template/error.html", {"error_message": f"Error calculating total expenditure: {str(e)}"})

    try:
        net_profit = total_fees_collected - total_expenditure
    except Exception as e:
        return render(request, "headmaster_template/error.html", {"error_message": f"Error calculating net profit: {str(e)}"})

    return render(request, "headmaster_template/home_content.html", {
        'total_students': total_students,
        'total_subjects': total_subjects,     
        'total_female_students': total_female_students,
        'total_male_students': total_male_students,
        'student_count': students_count,
        'total_fees_collected': total_fees_collected,
        'pending_payments_count': pending_payments_count,
        'total_expenditure': total_expenditure,
        'net_profit': net_profit,
        'total_school_fees': total_school_fees,
        'total_madrasatul_fees': total_madrasatul_fees,
        'total_transport_fees': total_transport_fees,
        'total_fee_payment': total_fee_payment
    })
    
    
@login_required
def student_general_attendance(request, student_id):
    try:
        if student_id:
            student_object = Students.objects.get(id=student_id)
        else:
            # Redirect to the student login page if not logged in
            return render(request, "headmaster_template/student_general_attendance.html")

        # Retrieve attendance data for the student


        class_attendance_present = StudentClassAttendance.objects.filter(student=student_object, status=True).count()
        class_attendance_absent = StudentClassAttendance.objects.filter(student=student_object, status=False).count()
        class_attendance_total = StudentClassAttendance.objects.filter(student=student_object).count()      
      
        total_subjects_taken = Subject.objects.all().count()

       

        # Pass data to the template for rendering
        return render(
            request,
            "headmaster_template/student_general_attendance.html",
            {
              
                "class_attendance_present": class_attendance_present,
                "class_attendance_absent": class_attendance_absent,
                "class_attendance_total": class_attendance_total,
                "total_subjects_taken": total_subjects_taken,
                                 
                "student": student_object,  # Renamed 'students' to 'student'
            },
        )

    except Students.DoesNotExist:
        messages.error(request, "Student does not exist.")
        return render(request, "headmaster_template/student_general_attendance.html")

    except Exception as e:
        messages.error(request, f"Error occurred: {e}")
        return render(request, "headmaster_template/student_general_attendance.html")
   

def get_financial_yearly_data(request):
    if request.method == 'GET' and 'year' in request.GET:
        selected_year = request.GET.get('year')
        
        # Get the branch of the currently logged-in staff
        current_branch = request.user.staffs.branch
        
        try:
            # Calculate total fees collected from all payment models for the selected year and branch
            school_fees_collected = SchoolFeesPayment.objects.filter(
                date_paid__year=selected_year, student__branch=current_branch
            ).aggregate(total=models.Sum('amount_paid'))['total'] or 0
            
            class_fees_collected = FeePayment.objects.filter(
                payment_date__year=selected_year, student__branch=current_branch
            ).aggregate(total=models.Sum('amount_paid'))['total'] or 0
            
            madrasatul_fees_collected = MadrasatulFeePayment.objects.filter(
                payment_date__year=selected_year, student__branch=current_branch
            ).aggregate(total=models.Sum('amount_paid'))['total'] or 0
            
            transport_fees_collected = TransportFeePayment.objects.filter(
                payment_date__year=selected_year, student__branch=current_branch
            ).aggregate(total=models.Sum('amount_paid'))['total'] or 0
            
            # Sum all the fees collected
            total_fees_collected = (school_fees_collected + class_fees_collected + madrasatul_fees_collected + transport_fees_collected)
            
            # Fetch expenditure for the current branch
            total_expenditure = Expenditure.objects.filter(
                date__year=selected_year, branch=current_branch
            ).aggregate(total=models.Sum('amount'))['total'] or 0
            
            # Prepare data to return as JSON
            data = {
                'income': total_fees_collected,
                'expenditure': total_expenditure,
            }
            return JsonResponse(data)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    # Handle invalid requests or missing parameters
    return JsonResponse({'error': 'Invalid request'}, status=400)



@require_GET
def get_school_fees_monthly_data(request):
    year = request.GET.get('year')
    
    # Get the branch of the currently logged-in staff
    current_branch = request.user.staffs.branch
    
    # Initialize monthly data dictionary
    monthly_data_dict = {month: {'income': 0, 'expenditure': 0} for month in range(1, 13)}

    # Fetch monthly income data from SchoolFeesPayment filtered by year and branch
    school_fees_data = SchoolFeesPayment.objects.filter(
        date_paid__year=year, student__branch=current_branch
    ).values('date_paid__month').annotate(
        income=Sum('amount_paid')
    ).order_by('date_paid__month')   

    # Combine all data into the monthly_data_dict
    for entry in school_fees_data:
        month = entry['date_paid__month']
        monthly_data_dict[month]['income'] += entry['income']    

    # Prepare response data
    response_data = []
    for month in range(1, 13):
        response_data.append({
            'month': month,
            'income': monthly_data_dict[month]['income'],       
        })

    return JsonResponse(response_data, safe=False)


@require_GET
def get_fee_payment_monthly_data(request):
    year = request.GET.get('year')
    
    # Get the branch of the currently logged-in staff
    current_branch = request.user.staffs.branch
    
    # Initialize monthly data dictionary
    monthly_data_dict = {month: {'income': 0, 'expenditure': 0} for month in range(1, 13)}

    # Fetch monthly income data from FeePayment filtered by year and branch
    fee_payment_data = FeePayment.objects.filter(
        payment_date__year=year, 
        student__branch=current_branch
    ).values('payment_date__month').annotate(
        income=Sum('amount_paid')
    ).order_by('payment_date__month')

    # Combine all data into the monthly_data_dict
    for entry in fee_payment_data:
        month = entry['payment_date__month']
        monthly_data_dict[month]['income'] += entry['income']

    # Prepare response data
    response_data = []
    for month in range(1, 13):
        response_data.append({
            'month': month,
            'income': monthly_data_dict[month]['income'],
        })

    return JsonResponse(response_data, safe=False)


@require_GET
def get_madrasatul_fee_payment_monthly_data(request):
    year = request.GET.get('year')
    
    # Get the branch of the currently logged-in staff
    current_branch = request.user.staffs.branch
    
    # Initialize monthly data dictionary
    monthly_data_dict = {month: {'income': 0} for month in range(1, 13)}

    # Fetch monthly income data from MadrasatulFeePayment filtered by year and branch
    madrasatul_fee_payment_data = MadrasatulFeePayment.objects.filter(
        payment_date__year=year, 
        student__branch=current_branch
    ).values('payment_date__month').annotate(
        income=Sum('amount_paid')
    ).order_by('payment_date__month')
    
    # Combine all data into the monthly_data_dict
    for entry in madrasatul_fee_payment_data:
        month = entry['payment_date__month']
        monthly_data_dict[month]['income'] += entry['income']

    # Prepare response data
    response_data = []
    for month in range(1, 13):
        response_data.append({
            'month': month,
            'income': monthly_data_dict[month]['income'],
        })

    return JsonResponse(response_data, safe=False)


@require_GET
def get_transport_fee_payment_monthly_data(request):
    year = request.GET.get('year')
    
    # Ensure the year is provided
    if not year:
        return JsonResponse({'error': 'Year parameter is required'}, status=400)
    
    # Get the branch of the currently logged-in staff
    staff = request.user.staffs
    current_branch = staff.branch

    # Initialize monthly data dictionary
    monthly_data_dict = {month: {'income': 0} for month in range(1, 13)}

    # Fetch monthly income data from TransportFeePayment filtered by year and branch
    transport_fee_payment_data = TransportFeePayment.objects.filter(
        payment_date__year=year,
        student__branch=current_branch
    ).values('payment_date__month').annotate(
        income=Sum('amount_paid')
    ).order_by('payment_date__month')

    # Combine all data into the monthly_data_dict
    for entry in transport_fee_payment_data:
        month = entry['payment_date__month']
        monthly_data_dict[month]['income'] += entry['income']

    # Prepare response data
    response_data = []
    for month in range(1, 13):
        response_data.append({
            'month': month,
            'income': monthly_data_dict[month]['income'],
        })

    return JsonResponse(response_data, safe=False)


@login_required
def filter_payments(request):
    installments  = SchoolFeesInstallment.objects.all()   
    class_levels = ClassLevel.objects.all()
    return render(request, "headmaster_template/filter_payments.html",{"installments":installments,"class_levels":class_levels}) 

@login_required
def student_payment_fetch(request):
    installments  = SchoolFeesInstallment.objects.all()   
    class_levels = ClassLevel.objects.all()
    return render(request, "headmaster_template/student_payment_fetch.html",{"installments":installments,"class_levels":class_levels}) 


@require_POST
def payment_fetch(request):
    if request.method == 'POST' and request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        # Extract form data
        class_level_id = request.POST.get('class_level')
        service = request.POST.get('services')
        payment_status = request.POST.get('payment_status')

        # Get class level
        try:
            class_level = ClassLevel.objects.get(id=class_level_id)
        except ClassLevel.DoesNotExist:
            return JsonResponse({'error': 'Invalid class level selected'})

        # Get the branch of the currently logged-in staff
        staff = request.user.staffs
        current_branch = staff.branch

        # Fetch all students of the selected class who are in the same branch as the staff
        students = Students.objects.filter(class_level_id=class_level_id, branch=current_branch)

        # Initialize student_results list
        student_results = []

        # Determine which payment model and template to use based on the selected service
        if service == 'Madrasa':
            payments_model = MadrasatulFeePayment
            template_name = 'headmaster_template/payment_fetch_madrasa_table.html'
            for student in students:
                payments = payments_model.objects.filter(
                    student=student,
                    payment_status=payment_status
                )
                for payment in payments:
                    student_results.append({
                        'student_name': student.full_name,
                        'class': student.class_level,
                        'amount_paid': payment.amount_paid,
                        'amount_remaining': payment.amount_remaining,
                        'payment_status': payment.payment_status,
                        'date_paid': payment.payment_date
                    })
        
        elif service == 'Transport':
            payments_model = TransportFeePayment
            template_name = 'headmaster_template/payment_fetch_transport_table.html'
            for student in students:
                payments = payments_model.objects.filter(
                    student=student,
                    payment_status=payment_status
                )
                for payment in payments:
                    student_results.append({
                        'student_name': student.full_name,
                        'class': student.class_level,
                        'amount_paid': payment.amount_paid,
                        'amount_remaining': payment.amount_remaining,
                        'payment_status': payment.payment_status,
                        'date_paid': payment.payment_date
                    })

        elif service in ['Registration', 'Boarding']:
            payments_model = FeePayment
            template_name = 'headmaster_template/payment_fetch_registration_boarding_table.html'
            
            for student in students:
                # Filter ClassFee based on the service type
                fee_type = 'Registration' if service == 'Registration' else 'Boarding'
                class_fees = ClassFee.objects.filter(
                    class_level=student.class_level,
                    fee_type=fee_type
                )
                
                for class_fee in class_fees:
                    payments = payments_model.objects.filter(
                        student=student,
                        class_fee=class_fee,
                        payment_status=payment_status
                    )
                    for payment in payments:
                        student_results.append({
                            'student_name': student.full_name,
                            'class': student.class_level,
                            'fee_type': class_fee.fee_type,
                            'amount_paid': payment.amount_paid,
                            'amount_remaining': payment.amount_remaining,
                            'payment_status': payment.payment_status,
                            'date_paid': payment.payment_date
                        })

        else:
            return JsonResponse({'error': 'Invalid service selected'})

        # Render the HTML template with the fetched data
        html_result = render_to_string(template_name, {'student_results': student_results})

        # Return the HTML result as JSON response
        return JsonResponse({'html_result': html_result})

    return JsonResponse({'error': 'Invalid request'})



CLASS_CHOICES = ClassLevel.objects.all()

@require_GET
def get_students_by_class(request):
    year = request.GET.get('year')
    
    # Get the branch of the currently logged-in staff
    staff = request.user.staffs
    current_branch = staff.branch

    # Initialize a dictionary with all classes and zero counts
    students_by_class = defaultdict(int)
    for class_name in CLASS_CHOICES:
        students_by_class[class_name] = 0

    # Fetch the student counts for the specified year, filtered by branch
    if year:
        student_data = Students.objects.filter(
            created_at__year=year,
            branch=current_branch
        ).values('class_level').annotate(
            student_count=Count('id')
        )
    else:
        student_data = Students.objects.filter(
            branch=current_branch
        ).values('class_level').annotate(
            student_count=Count('id')
        )

    # Update the counts in the dictionary
    for entry in student_data:
        class_name = entry['class_level']
        if class_name in students_by_class:
            students_by_class[class_name] = entry['student_count']

    data = {
        'labels': list(students_by_class.keys()),
        'data': list(students_by_class.values())
    }

    return JsonResponse(data)


@require_POST
def search_payments(request):
    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        # Extract form data
        class_level_id = request.POST.get('class_level')
        installment_id = request.POST.get('installment_id')
        payment_status = request.POST.get('payment_status')

        # Get the branch of the currently logged-in staff
        staff = request.user.staffs
        current_branch = staff.branch

        try:
            # Fetch the class level
            class_level = ClassLevel.objects.get(id=class_level_id)
        except ClassLevel.DoesNotExist:
            return JsonResponse({'error': 'Invalid class level selected'})

        # Fetch all students of the selected class and branch
        students = Students.objects.filter(
            class_level=class_level,
            branch=current_branch
        )

        # Create a list to store each student's payment results
        student_results = []

        # Iterate over each student to fetch their payment details
        for student in students:
            payments = SchoolFeesPayment.objects.filter(
                student=student,
                installment_id=installment_id,
                payment_status=payment_status
            )
            
            # Collect the payment details for each student
            for payment in payments:
                student_results.append({
                    'student_id': student.id,
                    'student_name': student.full_name,
                    'class': student.class_level,
                    'installment': payment.installment.installment_name,
                    'amount_paid': payment.amount_paid,
                    'amount_remaining': payment.amount_remaining,
                    'balance': payment.balance,
                    'payment_status': payment.payment_status,
                    'date_paid': payment.date_paid
                })

        # Render the HTML template with the fetched data
        html_result = render_to_string('headmaster_template/filter_payments_table.html', {'student_results': student_results})
        
        # Return the HTML result as JSON response
        return JsonResponse({'html_result': html_result})
    
    return JsonResponse({'error': 'Invalid request'})



def view_all_payments(request, student_id):
    payments = SchoolFeesPayment.objects.filter(student__id=student_id)
    student = Students.objects.get(id=student_id)
    return render(request, 'headmaster_template/view_all_payments.html', {'payments': payments,"student":student})


@require_GET
def get_expenditure_monthly_data(request):
    year = request.GET.get('year')
    
    # Get the branch of the currently logged-in staff
    staff = request.user.staffs
    current_branch = staff.branch

    # Initialize monthly data dictionary
    monthly_data_dict = {month: {'income': 0, 'expenditure': 0} for month in range(1, 13)}

    # Fetch monthly expenditure data filtered by branch
    expenditure_data = Expenditure.objects.filter(
        date__year=year,
        branch=current_branch  # Assuming Expenditure has a branch field
    ).values('date__month').annotate(
        expenditure=Sum('amount')
    ).order_by('date__month')

    # Populate the monthly_data_dict with the fetched expenditure data
    for entry in expenditure_data:
        month = entry['date__month']
        monthly_data_dict[month]['expenditure'] += entry['expenditure']

    # Prepare response data
    response_data = []
    for month in sorted(monthly_data_dict.keys()):
        response_data.append({
            'month': month,
            'income': monthly_data_dict[month]['income'],  # No income data for Expenditure
            'expenditure': monthly_data_dict[month]['expenditure'],
        })

    return JsonResponse(response_data, safe=False)


@require_GET
def get_classwise_fee_payment_data(request):
    try:
        year = request.GET.get('year')

        # Get the branch of the currently logged-in staff
        staff = request.user.staffs
        current_branch = staff.branch

        # Query to count number of distinct students per class who have made payments
        classwise_data = Students.objects.filter(
            schoolfeespayment__date_paid__year=year,
            branch=current_branch  # Filter by current branch
        ).values('class_level').annotate(
            fee_payments_count=Count('id', distinct=True)
        ).order_by('class_level')

        # Prepare response data
        response_data = []
        for entry in classwise_data:
            response_data.append({
                'class_name': entry['class_level'],
                'fee_payments_count': entry['fee_payments_count'],
            })

        return JsonResponse(response_data, safe=False)

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    

def student_list(request):
    # Get the branch of the currently logged-in staff
    staff = request.user.staffs
    current_branch = staff.branch

    # Fetch students filtered by the current branch
    students = Students.objects.filter(branch=current_branch)

    context = {
        'students': students
    }
    return render(request, 'headmaster_template/student_list.html', context)


def student_details(request, id):
    student = get_object_or_404(Students, id=id)
    return render(request, 'headmaster_template/student_details.html', {'student': student})

def school_fees_installments_list(request):
    installments = SchoolFeesInstallment.objects.all()
    context = {
        'installments': installments
    }
    return render(request, 'headmaster_template/school_fees_installments_list.html', context)


def fees_collected(request):
    # Get the branch of the currently logged-in staff
    staff = request.user.staffs
    current_branch = staff.branch

    # Calculate total fees collected for the current branch
    total_fees_collected = SchoolFeesPayment.objects.filter(
        student__branch=current_branch
    ).aggregate(total_collected=models.Sum('amount_paid'))['total_collected'] or 0

    # Get installments where amount_remaining is 0 for the current branch
    installments = SchoolFeesPayment.objects.filter(
        amount_remaining=0,
        student__branch=current_branch
    )

    context = {
        'installments': installments,
        'total_fees_collected': total_fees_collected
    }
    return render(request, 'headmaster_template/fees_collected.html', context)


def pending_payments(request):
    # Get the branch of the currently logged-in staff
    staff = request.user.staffs
    current_branch = staff.branch

    # Fetch pending payments for the current branch
    pending_payments = SchoolFeesPayment.objects.filter(
        payment_status='Incomplete',
        student__branch=current_branch
    )

    context = {
        'pending_payments': pending_payments
    }
    return render(request, 'headmaster_template/pending_payments.html', context)


def expenditure_list(request):
    # Get the branch of the currently logged-in staff
    staff = request.user.staffs
    current_branch = staff.branch

    # Fetch expenditures for the current branch
    expenditures = Expenditure.objects.filter(
        branch=current_branch
    )
    total_expenditure = Expenditure.objects.filter(
        branch=current_branch
    ).aggregate(total=models.Sum('amount'))['total'] or 0

    context = {
        'expenditures': expenditures,
        'total_expenditure': total_expenditure
    }
    return render(request, 'headmaster_template/expenditure.html', context)  


def download_excel_template(request, class_level_id):
    # Get the branch of the currently logged-in staff
    staff = request.user.staffs
    current_branch = staff.branch

    # Fetch the class level object
    class_level = get_object_or_404(ClassLevel, id=class_level_id)
    
    # Get students and subjects for the given class level and branch
    students = Students.objects.filter(class_level=class_level, branch=current_branch)
    subjects = Subject.objects.filter(class_level=class_level)

    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers for the Excel sheet
    headers = ['Student Registration Number', 'Student Name']
    for subject in subjects:
        headers.append(subject.subject_name.capitalize())  # Add subject names as headers
    ws.append(headers)

    # Add student registration numbers, names, and zero scores for each subject
    for student in students:
        row_data = [student.registration_number, student.full_name]  # Add student registration number and name
        for _ in subjects:
            row_data.append(0)  # Add a zero value for each subject
        ws.append(row_data)

    # Save the workbook in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Set response headers for Excel file download
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{class_level.class_name}_template.xlsx"'
    
    return response


@login_required
def manage_student(request):
    # Get the branch of the currently logged-in staff
    staff = request.user.staffs
    current_branch = staff.branch

    # Fetch students and exam types corresponding to the current branch
    students = Students.objects.filter(branch=current_branch)
    exam_types = ExamType.objects.all()

    # Render the template with the filtered data
    return render(request, 'headmaster_template/manage_student.html', {
        'students': students,
        'exam_types': exam_types,
    })
    
    
@login_required
def student_subject_wise_result_page(request,student_id): 
    student = Students.objects.get(id=student_id)
    exam_types = ExamType.objects.all()
    distinct_dates= Result.objects.order_by('date_of_exam').values_list('date_of_exam', flat=True).distinct()
    print(distinct_dates)
    return render(request, 'headmaster_template/subject_wise_results.html',
                  {
                      'student': student,
                      'exam_types': exam_types,
                      'distinct_dates': distinct_dates,
                   })

@login_required    
def students_wise_result_page(request):
    exam_types = ExamType.objects.all()
    class_levels = ClassLevel.objects.all()
    distinct_dates= Result.objects.order_by('date_of_exam').values_list('date_of_exam', flat=True).distinct()   
    return render(request, 'headmaster_template/student_results.html',
                  {                   
                      'class_levels': class_levels,
                      'exam_types': exam_types,
                      'distinct_dates': distinct_dates,
                   })
    
    
@login_required
def manage_exam_type(request):
    exam_types = ExamType.objects.all()
    return render(request, 'headmaster_template/manage_exam_types.html', {'exam_types': exam_types})

@login_required
def manage_subject(request):
    subjects = Subject.objects.all()
    return render(request, 'headmaster_template/manage_subject.html', {'subjects': subjects})



@csrf_exempt        
@require_POST
def save_exam_type(request):
    try:
        exam_type_id = request.POST.get('exam_type_id')
        name = request.POST.get('name')
        description = request.POST.get('description')
        
        if exam_type_id:
            # Editing existing inventory item
            exam_type = ExamType.objects.get(pk=exam_type_id)
            exam_type.name = name
            exam_type.description = description
                   
            exam_type.save()
        else:
            # Adding new inventory item
            exam_type = ExamType(
            name=name,
            description=description,
              
            )
            exam_type.save()

        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})     
    

def update_student_status(request):
    try:
        if request.method == 'POST':
            # Get the user_id and is_active values from POST data
            user_id = request.POST.get('user_id')
            is_active = request.POST.get('is_active')

            # Retrieve the staff object or return a 404 response if not found
            student = get_object_or_404(Students, id=user_id)

            # Toggle the is_active status based on the received value
            if is_active == '1':
                student.is_active = False
            elif is_active == '0':
                student.is_active = True
            else:
                messages.error(request, 'Invalid request')
                return JsonResponse("Invalid request") # Make sure 'manage_staffs' is the name of your staff list URL

            student.save()
            messages.success(request, 'Status updated successfully')
        else:
            messages.error(request, 'Invalid request method')
    except Exception as e:
        messages.error(request, f'An error occurred: {str(e)}')

    # Redirect back to the staff list page
    return redirect('manage_student')  # Make sure 'manage_staffs' is the name of your staff list URL



@login_required
def student_subject_wise_result(request):
    if request.method == 'POST' and request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        # Get the branch of the currently logged-in staff
        staff = request.user.staffs
        current_branch = staff.branch

        # Extract form data
        student_id = request.POST.get('student_id')
        exam_type_id = request.POST.get('exam_type_id')
        year = request.POST.get('year')

        # Get the student and ensure they belong to the current branch
        student = get_object_or_404(Students, id=student_id, branch=current_branch)

        # Ensure the student is from the correct branch
        form_i_students = Students.objects.filter(class_level=student.class_level, branch=current_branch)
        total_students = form_i_students.count()

        # Get the exam type and ensure it's valid
        exam_type = get_object_or_404(ExamType, id=exam_type_id)

        # Fetch results, exam info, position info, and exam metrics, all filtered by branch
        results = Result.objects.filter(student=student, exam_type=exam_type, date_of_exam=year)
        exam_info = StudentExamInfo.objects.filter(
            student=student,
            exam_type=exam_type,
            class_level=student.class_level
        ).first()

        position_info = StudentPositionInfo.objects.filter(
            student=student,
            exam_type=exam_type,
            class_level=student.class_level
        ).first()

        exam_metrics, created = ExamMetrics.objects.get_or_create(
            student=student,
            exam_type=exam_type,
            class_level=student.class_level,
        )
        total_marks = exam_metrics.total_marks
        average = exam_metrics.average
        grademetrics = exam_metrics.grade
        remark = exam_metrics.remark

        if exam_info:
            division = exam_info.division
            total_grade_points = exam_info.total_grade_points
        else:
            division = "Division Not Available"
            total_grade_points = "Total Grade Points Not Available"

        if position_info:
            position = position_info.position
        else:
            position = "Position Not Available"

        context = {
            'year': year,
            'remark': remark,
            'grademetrics': grademetrics,
            'average': average,
            'total_marks': total_marks,
            'student': student,
            'results': results,
            "students": student,
            'exam_type': exam_type,
            'position': position,  # Add position to the context
            'division': division,  # Add division to the context
            'total_students': total_students,  # Add total_students to the context
            'total_grade_points': total_grade_points,  # Add total_grade_points to the context
        }
        html_result = render_to_string('headmaster_template/result_table.html', context)
        return JsonResponse({'html_result': html_result})
    
    return JsonResponse({'error': 'Invalid request'})

    
    
def download_student_results_excel(request, student_id, exam_type_id, year):
    # Convert year string to a date or handle as needed
    year = year  # Adjust this as per your data format or handling needs

    # Get the student based on the provided student_id
    student = get_object_or_404(Students, id=student_id)
    exam_type = get_object_or_404(ExamType, id=exam_type_id)

    # Query results
    results = Result.objects.filter(student=student, exam_type=exam_type, date_of_exam=year)
    exam_info = StudentExamInfo.objects.filter(student=student, exam_type=exam_type, class_level=student.class_level).first()
    position_info = StudentPositionInfo.objects.filter(student=student, exam_type=exam_type, class_level=student.class_level).first()
    exam_metrics, created = ExamMetrics.objects.get_or_create(
                    student=student,
                    exam_type=exam_type,
                    class_level=student.class_level,
                )
    total_marks = exam_metrics.total_marks
    average = exam_metrics.average
    grademetrics = exam_metrics.grade
    remark = exam_metrics.remark    

    if exam_info:
        division = exam_info.division
        total_grade_points = exam_info.total_grade_points
    else:
        division = "Division Not Available"
        total_grade_points = "Total Grade Points Not Available"

    if position_info:
        position = position_info.position
    else:
        position = "Position Not Available"

    # Create Excel workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = f"{student.full_name} Results"

    # Headers
    ws.append(['Subject', 'Marks', 'Grade', 'Pass/Fail'])

    # Data rows
    for result in results:
        ws.append([
            result.subject.subject_name,
            result.marks,
            result.determine_grade(),  # Call the method to get the value
            result.determine_pass_fail()
        ])

    # Additional information
    ws.append(['Position', position])
    ws.append(['Total Grade Points', total_grade_points])
    ws.append(['Division', division])
    ws.append(['Total Marks', total_marks])
    ws.append(['Average', average])
    ws.append(['Grade', grademetrics])
    ws.append(['Remark', remark])

    # Adjust column widths and alignment
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
        ws.column_dimensions[column].alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        ws.cell(row=1, column=col[0].column).font = Font(bold=True)

    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={student.full_name}_results.xlsx'
    wb.save(response)
    return response    


@csrf_exempt 
def save_marks_view(request):
    if request.method == 'POST':
        try:
            # Retrieve data from the POST request
            result_id = request.POST.get('resultId')
            marks = request.POST.get('marks')

            # Update the marks for the corresponding result
            result = Result.objects.get(id=result_id)
            result.marks = marks
            result.save()

            # Return success response
            return JsonResponse({'success': True, 'message': 'Marks updated successfully.'})

        except Exception as e:
            # Return error response
            return JsonResponse({'success': False, 'message': str(e)})
    else:
        # Return error response for unsupported request method
        return JsonResponse({'success': False, 'message': 'Only POST requests are allowed for this endpoint.'})
    
    
@csrf_exempt  # Use this decorator if CSRF protection is not needed
def delete_result_endpoint(request):
    if request.method == 'POST':
        try:
            # Extract the result ID from the request
            result_id = request.POST.get('resultId')

            # Perform the deletion operation, assuming Result is the model name
            result = Result.objects.get(id=result_id)
            result.delete()

            # Return success response
            return JsonResponse({'success': True, 'message': 'Result deleted successfully.'})

        except Exception as e:
            # Return error response
            return JsonResponse({'success': False, 'message': str(e)})

    else:
        # Return error response for non-POST requests
        return JsonResponse({'success': False, 'message': 'Only POST requests are allowed for this endpoint.'})
        

@login_required
def student_results_view(request):
    if request.method == 'POST' and request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        # Get the branch of the currently logged-in staff
        staff = request.user.staffs
        current_branch = staff.branch

        # Extract form data from POST request
        class_level_id = request.POST.get('class_level')
        exam_type_id = request.POST.get('exam_type_id')
        date_of_exam = request.POST.get('date_of_exam')
        
        # Fetch the class level object
        class_level = get_object_or_404(ClassLevel, id=class_level_id)
        
        # Fetch all students for the given class level and branch
        students = Students.objects.filter(class_level=class_level, branch=current_branch)
        
        # Fetch all subjects for the given class level
        subjects = Subject.objects.filter(class_level=class_level)
        
        # List to store each student's results and metrics
        student_results = []
        
        # Iterate over each student
        for student in students:
            # Fetch results for the current student filtered by exam type, date, and class level
            results = Result.objects.filter(
                student=student,
                exam_type_id=exam_type_id,
                date_of_exam=date_of_exam,
                class_level=class_level,
            )
            
            # Initialize variables for position, division, total grade points, total marks, average, grade, and remark
            position = None
            division = None
            total_grade_points = None
            total_marks = None
            average = None
            grademetrics = None
            remark = None
            
            # Fetch or calculate ExamMetrics, PositionInfo, and StudentExamInfo
            exam_metrics, created = ExamMetrics.objects.get_or_create(
                student=student,
                exam_type_id=exam_type_id,
                class_level=class_level,
            )
            total_marks = exam_metrics.total_marks
            average = exam_metrics.average
            grademetrics = exam_metrics.grade
            remark = exam_metrics.remark

            position_info = StudentPositionInfo.objects.filter(
                student=student,
                exam_type_id=exam_type_id,
                class_level=class_level
            ).first()
            position = position_info.position if position_info else "Position Not Available"
            
            student_exam_info = StudentExamInfo.objects.filter(
                student=student,
                exam_type_id=exam_type_id,
                class_level=class_level
            ).first()
            if student_exam_info:
                division = student_exam_info.division
                total_grade_points = student_exam_info.total_grade_points
            else:
                division = "Division Not Available"
                total_grade_points = "Total Grade Points Not Available"

            # Create a dictionary to store subject-wise results for the current student
            student_subject_results = {}
            
            # Iterate over each subject
            for subject in subjects:
                # Check if there is a result for the current subject and student
                result = results.filter(subject=subject).first()
                if result:
                    # If result exists, store the grade
                    grade = result.determine_grade()
                else:
                    # If result does not exist, mark as 'Not Assigned'
                    grade = ''
                
                # Store subject-wise grades in the dictionary
                student_subject_results[subject.subject_name] = grade
            
            # Append student results and metrics to the list
            student_results.append({
                'student': student,
                'subjects': subjects,
                'total_marks': total_marks,
                'average': average,
                'grademetrics': grademetrics,
                'position': position,
                'division': division,
                'total_grade_points': total_grade_points,
                'remark': remark,
                'student_subject_results': student_subject_results,
            })
        
        # Render the HTML template with the fetched data
        html_result = render_to_string('headmaster_template/student_results_table.html', {'student_results': student_results, 'subjects': subjects})
        
        # Return the HTML result as JSON response
        return JsonResponse({'html_result': html_result})
    
    return JsonResponse({'error': 'Invalid request'})
    

@login_required
def fetch_students_per_class(request):
    # Get the branch of the currently logged-in staff
    staff = request.user.staffs
    current_branch = staff.branch

    # Get the total number of students for each class level in the current branch
    class_levels = Students.objects.filter(branch=current_branch).values('class_level__class_name').annotate(
        total_students=Count('id')
    )
    
    # Prepare data for JSON response
    labels = [item['class_level__class_name'] for item in class_levels]
    data = [item['total_students'] for item in class_levels]
    
    return JsonResponse({
        'labels': labels,
        'data': data
    })


@login_required
def exam_type_list(request):
    exam_types = ExamType.objects.all()
    return render(request, 'headmaster_template/exam_type_list.html', {'exam_types': exam_types})

@login_required
def exam_type_to_view_class(request, exam_type_id):
    try:
        # Retrieve the exam type object
        exam_type = ExamType.objects.get(pk=exam_type_id)
        class_levels = ClassLevel.objects.all()
        # Pass the exam type object to the template
        return render(request, 'headmaster_template/class_wise_results.html', {'exam_type': exam_type,"class_levels":class_levels})
    except ExamType.DoesNotExist:
        # Handle the case where the exam type does not exist
        return redirect('exam_type_list') 



@login_required
def view_student_to_add_result(request, exam_type_id, class_level_id):
    # Get the branch of the currently logged-in staff
    staff = request.user.staffs
    current_branch = staff.branch
    
    class_level = get_object_or_404(ClassLevel, id=class_level_id)
    exam_type = get_object_or_404(ExamType, pk=exam_type_id)
    subjects = Subject.objects.filter(class_level=class_level)
    
    # Fetch students from the specific class level and branch of the logged-in staff
    students = Students.objects.filter(class_level=class_level, branch=current_branch)
    
    student_results = []

    for student in students:
        results = Result.objects.filter(
            student=student,
            exam_type=exam_type,
            class_level=class_level,
        )
        
        position, division, total_grade_points = None, None, None
        total_marks, average, grademetrics, remark = None, None, None, None

        # Fetch or calculate metrics
        if results.exists():
            position_info, _ = StudentPositionInfo.objects.get_or_create(
                student=student,
                exam_type=exam_type,
                class_level=class_level,
            )
            position = position_info.position

            student_exam_info, _ = StudentExamInfo.objects.get_or_create(
                student=student,
                exam_type=exam_type,
                class_level=class_level,
            )
            division = student_exam_info.division
            total_grade_points = student_exam_info.total_grade_points

            exam_metrics, _ = ExamMetrics.objects.get_or_create(
                student=student,
                exam_type=exam_type,
                class_level=class_level,
            )
            total_marks = exam_metrics.total_marks
            average = exam_metrics.average
            grademetrics = exam_metrics.grade
            remark = exam_metrics.remark

        # Prepare subject results
        student_subject_results = {}
        for subject in subjects:
            result = results.filter(subject=subject).first()
            if result:
                grade = result.determine_grade()
            else:
                grade = ''
            
            student_subject_results[subject.subject_name] = grade

        student_results.append({
            'student': student,
            'total_marks': total_marks,
            'average': average,
            'grademetrics': grademetrics,
            'position': position,
            'division': division,
            'total_grade_points': total_grade_points,
            'remark': remark,
            'student_subject_results': student_subject_results,
        })

    return render(request, 'headmaster_template/class_wise_students_list.html', {
        'student_results': student_results,
        'class_name': class_level.class_name,
        'className': class_level.class_name,
        'class_level': class_level,
        'subjects': subjects,
        'exam_type': exam_type,
    })




@login_required
def display_results(request):
    # Get the branch of the currently logged-in staff
    staff = request.user.staffs
    current_branch = staff.branch
    
    # Filter students by the current staff's branch
    students = Students.objects.filter(branch=current_branch)
    
    # Filter results by the current staff's branch and include related students
    results = SujbectWiseResults.objects.filter(student__in=students)
    
    exam_types = ExamType.objects.all()
    
    return render(request, 'headmaster_template/exam_results.html', {
        'results': results,
        'students': students,
        'exam_types': exam_types,
    })


@login_required
def manage_staff(request):     
    staffs=Staffs.objects.all()  
    subjects=Subject.objects.all()  
    return render(request,"headmaster_template/manage_staff.html",{
        "staffs":staffs,
        "subjects":subjects,
        })      



@csrf_exempt
def delete_staff(request):
    if request.method == 'POST':
        staff_id = request.POST.get('staff_id')
        
        try:
            staff = get_object_or_404(Staffs, id=staff_id)
            custom_user = staff.admin            
            # Delete the staff record
            staff.delete()
            
            # Delete the associated CustomUser record
            custom_user.delete()
            
            return JsonResponse({'status': 'success', 'message': 'Staff record and associated user deleted successfully.'})
        except IntegrityError as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': 'An error occurred while deleting the staff record.'}, status=500)
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method.'}, status=400)


@login_required
def update_staff_status(request):
    try:
        if request.method == 'POST':
            # Get the user_id and is_active values from POST data
            user_id = request.POST.get('user_id')
            is_active = request.POST.get('is_active')

            # Retrieve the staff object or return a 404 response if not found
            staff = get_object_or_404(CustomUser, id=user_id)

            # Toggle the is_active status based on the received value
            if is_active == '1':
                staff.is_active = False
                messages.success(request, f'{staff.username} has been deactivated.')
            elif is_active == '0':
                staff.is_active = True
                messages.success(request, f'{staff.username} has been activated.')

                # Send activation email to the user
                login_url = request.build_absolute_uri(reverse('login'))
                email_content = (
                    f"Dear {staff.first_name} {staff.last_name},\n\n"
                    f"MRISHO HAMISI is pleased to inform you that your account has been activated successfully. "
                    f"You can now log in to your account using the following link:\n\n"
                    f"{login_url}\n\n"
                    f"Your login details are as follows:\n"
                    f"Email: {staff.email}\n"                
                    f"If you have any questions or need further assistance, please do not hesitate to contact the administrator.\n\n"
                    f"Best regards,\n"
                    f"ZAMZAM Team"
                )
                send_mail(
                    'Account Activation Notice',
                    email_content,
                    'admin@example.com',
                    [staff.email],
                    fail_silently=False,
                )
            else:
                messages.error(request, 'Invalid request')
                return redirect('headmaster_manage_staff')  

            staff.save()
        else:
            messages.error(request, 'Invalid request method')
    except Exception as e:
        messages.error(request, f'An error occurred: {str(e)}')

    # Redirect back to the staff list page
    return redirect('headmaster_manage_staff') 


@csrf_exempt
def staff_feedback_message_replied(request):
    feedback_id = request.POST.get("id")
    feedback_message = request.POST.get("message")
    try:    
      feedback = FeedBackStaff.objects.get(id=feedback_id)
      feedback.feedback_reply = feedback_message
      feedback.save() 
      return HttpResponse(True)
    except:
       return HttpResponse(False)           
 


@login_required
def staff_leave_view(request):
    leaves = LeaveReportStaffs.objects.all()
    return render(request,"headmaster_template/staff_leave_view.html",{"leaves":leaves})

def staff_approve_leave(request,leave_id):
    leave = LeaveReportStaffs.objects.get(id=leave_id)
    leave.leave_status = 1
    leave.save()
    return HttpResponseRedirect(reverse("headmaster_staff_leave_view"))


def staff_disapprove_leave(request,leave_id):
    leave = LeaveReportStaffs.objects.get(id=leave_id)
    leave.leave_status = 2
    leave.save()
    return HttpResponseRedirect(reverse("headmaster_staff_leave_view"))

@login_required
def view_attendance(request):     
     return render(request,"headmaster_template/headmaster_view_attendance.html",{})


@login_required 
def staff_feedback_message(request):
    feedbacks = FeedBackStaff.objects.all()
    return render(request,"headmaster_template/staff_feedback_message.html",{"feedbacks":feedbacks})



@csrf_exempt
def check_email_exist(request):
    email = request.POST.get("email")
    user_object = CustomUser.objects.filter(email=email).exists()
    if user_object:
        return HttpResponse(True)
    
    else:
        return HttpResponse(False)
    
    
@csrf_exempt
def check_username_exist(request):
    username = request.POST.get("username")
    user_object = CustomUser.objects.filter(username=username).exists()
    if user_object:
        return HttpResponse(True)
    
    else:
        return HttpResponse(False) 
   
@login_required   
def view_class_attendance(request):
    admin = request.user
    attendances=ClassAttendance.objects.all()  
    class_levels = ClassLevel.objects.all()
    return render(request, "headmaster_template/admin_view_class_attendance.html", {          
         "admin": admin,
        "attendances": attendances,
        "class_levels": class_levels,
    })

 
    
@csrf_exempt
@login_required
def get_class_student_attendance_data(request):
    if request.method == 'POST':
        date_id = request.POST.get("date")
        class_level = request.POST.get("class_level")

        # Get the branch of the currently logged-in staff
        staff = request.user.staffs
        current_branch = staff.branch

        # Filter students by the current staff's branch and class level
        students = Students.objects.filter(class_level=class_level, branch=current_branch)

        # Filter attendance by the current staff's branch
        attendance = ClassAttendance.objects.filter(id=date_id)

        # Fetch attendance data for the given date and class
        attendance_data = StudentClassAttendance.objects.filter(
            attendance__in=attendance,
            student__in=students
        )

        # Serialize student data
        list_data = []
        for record in attendance_data:
            data_small = {
                "id": record.student.id,
                "name": record.student.full_name,
                "status": record.status
            }
            list_data.append(data_small)
        
        # Return the serialized data as JSON response
        return JsonResponse(list_data, safe=False)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=400)


@login_required
def update_profile_picture(request):
    if request.method == 'POST':
        try:
            # Get the user's profile
            staff = request.user.staffs
            # Accessing the uploaded picture from the form
            new_picture = request.FILES.get('profile_picture')  
            
            if new_picture:
                # Check if the uploaded file is an image
                if not new_picture.content_type.startswith('image'):
                    raise ValidationError("Invalid image file type. Please upload a valid image.")

                # Check if the file size is within limit (2MB)
                if new_picture.size > 2 * 1024 * 1024:  # 2MB in bytes
                    raise ValidationError("File size exceeds the limit. Please upload a file smaller than 2MB.")               
              
                
                if staff:
                    # Update the profile picture
                    staff.profile_pic = new_picture
                    
                    # Save the changes
                    staff.save()
                    
                    # Redirect to a success page or back to the profile page
                    return render(request, 'headmaster_template/update_profile_picture.html')
                else:
                    raise ValidationError("User profile does not exist.")
            else:
                raise ValidationError("No profile picture uploaded.")
        except ValidationError as e:
            # Handle validation errors (e.g., invalid file type or size)
            messages.error(request, str(e))
        except Exception as e:
            # Handle any other unexpected errors
            messages.error(request, str(e))
            # You may want to log this error for debugging purposes
            
    return render(request, 'headmaster_template/update_profile_picture.html') 


def add_staff(request):
    try:
        if request.method == 'POST':
            form = AddStaffForm(request.POST, request.FILES)
            if form.is_valid():
                form.save()
                action = request.POST.get('action')
                if action == 'save':
                    return redirect('headmaster_manage_staff')
                elif action == 'save_and_continue':
                    return redirect('headmaster_add_staff')
        else:
            form = AddStaffForm()

        return render(request, 'headmaster_template/add_staff.html', {'form': form})

    except Exception as e:
        messages.error(request, f'An error occurred: {e}')
        return redirect('headmaster_manage_staff')

def edit_staff(request, pk):
    try:
        staff = get_object_or_404(Staffs, pk=pk)
        initial_data = {
            'email': staff.admin.email,           
            'first_name': staff.admin.first_name,
            'last_name': staff.admin.last_name,
        }

        if request.method == 'POST':
            form = EditStaffForm(request.POST, request.FILES, instance=staff)
            if form.is_valid():
                form.save()
                action = request.POST.get('action')
                if action == 'save':
                    return redirect('headmaster_manage_staff')
                elif action == 'save_and_continue':
                    return redirect('headmaster_add_staff')
        else:
            form = EditStaffForm(instance=staff, initial=initial_data)

        return render(request, 'headmaster_template/edit_staff.html', {'form': form, 'staff': staff})

    except Http404:
        messages.error(request, 'Staff not found.')
        return redirect('headmaster_manage_staff')

    except Exception as e:
        messages.error(request, f'An error occurred: {e}')
        return redirect('headmaster_manage_staff')



def add_or_edit_student(request, pk=None):
    if pk:
        student = get_object_or_404(Students, pk=pk)
    else:
        student = Students()

    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES, instance=student)
        if form.is_valid():
            form.save()
            action = request.POST.get('action')
            if action == 'save':
                return redirect('headmaster_manage_student')
            elif action == 'save_and_continue':
                return redirect('headmaster_add_student')
    else:
        form = StudentForm(instance=student)

    return render(request, 'headmaster_template/add_student.html', {'form': form, 'student': student})

class ClassLevelCreateUpdateView(View):
    def get(self, request, *args, **kwargs):
        # Check if there's a 'pk' in the URL kwargs, indicating edit mode
        class_level = None
        if 'pk' in kwargs:
            class_level = get_object_or_404(ClassLevel, pk=kwargs['pk'])
            form = ClassLevelForm(instance=class_level)
        else:
            form = ClassLevelForm()

        return render(request, 'headmaster_template/add_classlevel_form.html', {'form': form, 'is_edit': class_level is not None})
    
    def post(self, request, *args, **kwargs):
        # Check if 'pk' is in the URL kwargs, to handle form submission for edit mode
        class_level = None
        if 'pk' in kwargs:
            class_level = get_object_or_404(ClassLevel, pk=kwargs['pk'])
            form = ClassLevelForm(request.POST, instance=class_level)
        else:
            form = ClassLevelForm(request.POST)

        if form.is_valid():
            form.save()

            # Determine which button was clicked
            action = request.POST.get('action')
            if action == 'save':
                return redirect('headmaster_classlevel_list')
            elif action == 'save_and_continue':
                return redirect('headmaster_classlevel_create')

        # Re-render the form with validation errors
        return render(request, 'headmaster_template/add_classlevel_form.html', {'form': form, 'is_edit': class_level is not None})

class ClassLevelListView(ListView):
    model = ClassLevel
    template_name = 'headmaster_template/classlevel_list.html'
    context_object_name = 'class_levels'
    paginate_by = 20  # Optional: If you want to paginate the results

    def get_queryset(self):
        # Customize the queryset if needed, otherwise it will return all records
        return ClassLevel.objects.all()
    

    
    
class ClassLevelDeleteView(View):
    def post(self, request, pk, *args, **kwargs):
        class_level = get_object_or_404(ClassLevel, pk=pk)
        class_level.delete()
        return redirect('headmaster_classlevel_list')  
    
class SubjectDeleteView(View):
    def post(self, request, pk, *args, **kwargs):
        class_level = get_object_or_404(Subject, pk=pk)
        class_level.delete()
        return redirect('headmaster_manage_subject')  
class SchoolFeesInstallmentDeleteView(View):
    def post(self, request, pk, *args, **kwargs):
        class_level = get_object_or_404(SchoolFeesInstallment, pk=pk)
        class_level.delete()
        return redirect('headmaster_installment_list')  
    
class SubjectCreateUpdateView(View):
    def get(self, request, *args, **kwargs):
        subject = None
        if 'pk' in kwargs:
            subject = get_object_or_404(Subject, pk=kwargs['pk'])
            form = SubjectForm(instance=subject)
        else:
            form = SubjectForm()

        return render(request, 'headmaster_template/add_subject_form.html', {'form': form, 'is_edit': subject is not None})
    
    def post(self, request, *args, **kwargs):
        subject = None
        if 'pk' in kwargs:
            subject = get_object_or_404(Subject, pk=kwargs['pk'])
            form = SubjectForm(request.POST, instance=subject)
        else:
            form = SubjectForm(request.POST)

        if form.is_valid():
            form.save()

            action = request.POST.get('action')
            if action == 'save':
                return redirect('headmaster_manage_subject')
            elif action == 'save_and_continue':
                return redirect('headmaster_subject_create')

        return render(request, 'headmaster_template/add_subject_form.html', {'form': form, 'is_edit': subject is not None})    
    
class SchoolFeesInstallmentListView(ListView):
    model = SchoolFeesInstallment
    template_name = 'headmaster_template/school_fees_installments_list.html'
    context_object_name = 'installments'
    paginate_by = 20  # Optional: If you want to paginate the results

    def get_queryset(self):
        # Customize the queryset if needed, otherwise it will return all records
        return SchoolFeesInstallment.objects.all()
        
class SchoolFeesInstallmentCreateView(View):
    def get(self, request, pk=None, *args, **kwargs):
        if pk:
            installment = get_object_or_404(SchoolFeesInstallment, pk=pk)
            form = SchoolFeesInstallmentForm(instance=installment)
        else:
            form = SchoolFeesInstallmentForm()
        return render(request, 'headmaster_template/add_installment_form.html', {'form': form})
    
    def post(self, request, pk=None, *args, **kwargs):
        if pk:
            installment = get_object_or_404(SchoolFeesInstallment, pk=pk)
            form = SchoolFeesInstallmentForm(request.POST, instance=installment)
        else:
            form = SchoolFeesInstallmentForm(request.POST)
        
        if form.is_valid():
            form.save()
            action = request.POST.get('action')
            if action == 'save':
                messages.success(request, 'Installment saved successfully!')
                return redirect('headmaster_installment_list')
            elif action == 'save_and_continue':
                messages.success(request, 'Installment saved successfully! You can add another one.')
                return redirect('headmaster_installment_create')
        else:
            messages.error(request, 'Please correct the errors below.')

        return render(request, 'headmaster_template/add_installment_form.html', {'form': form})     
    
    
class FeeStructureListView(ListView):
    model = FeeStructure
    template_name = 'headmaster_template/school_fees_structure_list.html'
    context_object_name = 'fee_structures'
    paginate_by = 20  # Optional: If you want to paginate the results

    def get_queryset(self):
        # Customize the queryset if needed, otherwise it will return all records
        return FeeStructure.objects.all()   
    
class FeeStructureDeleteView(View):
    def post(self, request, pk, *args, **kwargs):
        class_level = get_object_or_404(FeeStructure, pk=pk)
        class_level.delete()
        return redirect('headmaster_fee_structure_list')       
    
class FeeStructureCreateView(View):
    def get(self, request, pk=None, *args, **kwargs):
        if pk:
            fee_structure = get_object_or_404(FeeStructure, pk=pk)
            form = FeeStructureForm(instance=fee_structure)
        else:
            form = FeeStructureForm()

        return render(request, 'headmaster_template/add_fee_structure_form.html', {'form': form})

    def post(self, request, pk=None, *args, **kwargs):
        if pk:
            fee_structure = get_object_or_404(FeeStructure, pk=pk)
            form = FeeStructureForm(request.POST, instance=fee_structure)
        else:
            form = FeeStructureForm(request.POST)

        if form.is_valid():
            form.save()
            action = request.POST.get('action')
            if action == 'save':
                return redirect('headmaster_fee_structure_list')
            elif action == 'save_and_continue':
                return redirect('headmaster_fee_structure_create')

        return render(request, 'headmaster_template/add_fee_structure_form.html', {'form': form})   
    
    
class ExpenditureListView(View):
    def get(self, request, *args, **kwargs):
        expenditures = Expenditure.objects.all()
        return render(request, 'headmaster_template/expenditure_list.html', {'expenditures': expenditures})   
    
class ExpenditureDeleteView(View):
    def post(self, request, pk, *args, **kwargs):
        expenditure = get_object_or_404(Expenditure, pk=pk)
        expenditure.delete()
        return redirect('headmaster_expenditure_list')       
    
class ExpenditureCreateView(View):
    def get(self, request, pk=None, *args, **kwargs):
        if pk:
            expenditure = get_object_or_404(Expenditure, pk=pk)
            form = ExpenditureForm(instance=expenditure)
        else:
            form = ExpenditureForm()

        return render(request, 'headmaster_template/add_expenditure_form.html', {'form': form})

    def post(self, request, pk=None, *args, **kwargs):
        if pk:
            expenditure = get_object_or_404(Expenditure, pk=pk)
            form = ExpenditureForm(request.POST, instance=expenditure)
        else:
            form = ExpenditureForm(request.POST)

        if form.is_valid():
            form.save()
            action = request.POST.get('action')
            if action == 'save':
                return redirect('headmaster_expenditure_list')  # Adjust the redirect to match your URL name
            elif action == 'save_and_continue':
                return redirect('headmaster_add_expenditure')  # Adjust the redirect to match your URL name

        return render(request, 'headmaster_template/add_expenditure_form.html', {'form': form}) 
    
    
class ClassFeeListView(ListView):
    model = ClassFee
    template_name = 'headmaster_template/class_fee_list.html'
    context_object_name = 'class_fees'
    paginate_by = 10  # Adjust the number of items per page if needed

    def get_queryset(self):
        return ClassFee.objects.all().order_by('class_level', 'fee_type')
    
class ClassFeeDeleteView(View):
    def post(self, request, pk, *args, **kwargs):
        class_fee = get_object_or_404(ClassFee, pk=pk)
        class_fee.delete()
        return redirect('headmaster_class_fee_list')      
        
class ClassFeeCreateView(View):
    def get(self, request, pk=None, *args, **kwargs):
        if pk:
            class_fee = get_object_or_404(ClassFee, pk=pk)
            form = ClassFeeForm(instance=class_fee)
        else:
            form = ClassFeeForm()

        return render(request, 'headmaster_template/add_class_fee_form.html', {'form': form})

    def post(self, request, pk=None, *args, **kwargs):
        if pk:
            class_fee = get_object_or_404(ClassFee, pk=pk)
            form = ClassFeeForm(request.POST, instance=class_fee)
        else:
            form = ClassFeeForm(request.POST)

        if form.is_valid():
            form.save()
            action = request.POST.get('action')
            if action == 'save':
                return redirect('headmaster_class_fee_list')  # Adjust the redirect to match your URL name
            elif action == 'save_and_continue':
                return redirect('headmaster_add_class_fee')  # Adjust the redirect to match your URL name

        return render(request, 'headmaster_template/add_class_fee_form.html', {'form': form})    
    


class MadrasatulFeeListView(ListView):
    model = MadrasatulFee
    template_name = 'headmaster_template/madrasatul_fee_list.html'
    context_object_name = 'fees'
    paginate_by = 10  # Optional: to paginate the list if you have many records

    def get_queryset(self):
        return MadrasatulFee.objects.select_related('class_level').order_by('class_level', 'fee_amount')
    
class MadrasatulFeeDeleteView(View):
    def post(self, request, pk, *args, **kwargs):
        class_fee = get_object_or_404(MadrasatulFee, pk=pk)
        class_fee.delete()
        return redirect('headmaster_madrasatul_fee_list')        
    
class MadrasatulFeeCreateView(View):
    def get(self, request, pk=None, *args, **kwargs):
        if pk:
            fee = get_object_or_404(MadrasatulFee, pk=pk)
            form = MadrasatulFeeForm(instance=fee)
        else:
            form = MadrasatulFeeForm()

        return render(request, 'headmaster_template/add_madrasatul_fee_form.html', {'form': form})

    def post(self, request, pk=None, *args, **kwargs):
        if pk:
            fee = get_object_or_404(MadrasatulFee, pk=pk)
            form = MadrasatulFeeForm(request.POST, instance=fee)
        else:
            form = MadrasatulFeeForm(request.POST)

        if form.is_valid():
            form.save()
            action = request.POST.get('action')
            if action == 'save':
                messages.success(request, 'Madrasatul Fee saved successfully!')
                return redirect('headmaster_madrasatul_fee_list')
            elif action == 'save_and_continue':
                messages.success(request, 'Madrasatul Fee saved successfully! You can add another fee.')
                return redirect('headmaster_madrasatul_fee_create')
        else:
            messages.error(request, 'Please correct the errors below.')

        return render(request, 'headmaster_template/add_madrasatul_fee_form.html', {'form': form})    
    
    
class TransportFeeListView(ListView):
    model = TransportFee
    template_name = 'headmaster_template/transport_fee_list.html'
    context_object_name = 'transport_fees'
    paginate_by = 10  # Optional: to paginate the list if you have many records


    
class TransportFeeDeleteView(View):
    def post(self, request, pk, *args, **kwargs):
        class_fee = get_object_or_404(TransportFee, pk=pk)
        class_fee.delete()
        return redirect('headmaster_transport_fee_list')          


class TransportFeeCreateUpdateView(View):
    def get(self, request, pk=None, *args, **kwargs):
        if pk:
            transport_fee = get_object_or_404(TransportFee, pk=pk)
            form = TransportFeeForm(instance=transport_fee)
            context = {
                'form': form,
                'is_edit': True
            }
        else:
            form = TransportFeeForm()
            context = {
                'form': form,
                'is_edit': False
            }
        return render(request, 'headmaster_template/add_transport_fee_form.html', context)
    
    def post(self, request, pk=None, *args, **kwargs):
        if pk:
            transport_fee = get_object_or_404(TransportFee, pk=pk)
            form = TransportFeeForm(request.POST, instance=transport_fee)
        else:
            form = TransportFeeForm(request.POST)

        if form.is_valid():
            form.save()
            action = request.POST.get('action')
            if action == 'save':
                messages.success(request, 'Transport Fee has been saved successfully!')
                return redirect('headmaster_transport_fee_list')
            elif action == 'save_and_continue':
                messages.success(request, 'Transport Fee has been saved! You can add another one.')
                return redirect('headmaster_transport_fee_create')

        context = {
            'form': form,
            'is_edit': pk is not None
        }
        return render(request, 'headmaster_template/add_transport_fee_form.html', context)
    

class FeePaymentListView(ListView):
    model = FeePayment
    template_name = 'headmaster_template/fee_payment_list.html'
    context_object_name = 'fee_payments'
    paginate_by = 10  # Adjust the number as needed

    def get_queryset(self):
        return FeePayment.objects.select_related('student', 'class_fee').order_by('-payment_date')
    

class FeePaymentDeleteView(View):
    def post(self, request, pk, *args, **kwargs):
        fee_payment = get_object_or_404(FeePayment, pk=pk)
        fee_payment.delete()
        return redirect('headmaster_fee_payment_list')          
    
@method_decorator(login_required, name='dispatch')    
class FeePaymentCreateUpdateView(View):
    def get(self, request, pk=None, *args, **kwargs):
        try:
            current_staff = Staffs.objects.get(admin=request.user)
            staff_branch = current_staff.branch
        except Staffs.DoesNotExist:
            return redirect('some_error_page')  # Replace with an appropriate error page

        # Get students from the branch of the current staff
        branch_students = Students.objects.filter(branch=staff_branch)

        if pk:
            fee_payment = get_object_or_404(FeePayment, pk=pk)
            form = FeePaymentForm(instance=fee_payment, branch_students=branch_students)
            context = {
                'form': form,
                'is_edit': True
            }
        else:
            form = FeePaymentForm(branch_students=branch_students)
            context = {
                'form': form,
                'is_edit': False
            }
        
        return render(request, 'headmaster_template/add_fee_payment_form.html', context)

    def post(self, request, pk=None, *args, **kwargs):
        try:
            current_staff = Staffs.objects.get(admin=request.user)
            staff_branch = current_staff.branch
        except Staffs.DoesNotExist:
            return redirect('some_error_page')  # Replace with an appropriate error page

        # Get students from the branch of the current staff
        branch_students = Students.objects.filter(branch=staff_branch)

        if pk:
            fee_payment = get_object_or_404(FeePayment, pk=pk)
            form = FeePaymentForm(request.POST, instance=fee_payment, branch_students=branch_students)
        else:
            form = FeePaymentForm(request.POST, branch_students=branch_students)

        if form.is_valid():
            fee_payment = form.save()
            action = request.POST.get('action')
            if action == 'save':
                messages.success(request, 'Fee Payment has been saved successfully!')
                return redirect('headmaster_fee_payment_list')
            elif action == 'save_and_continue':
                messages.success(request, 'Fee Payment has been saved! You can add another one.')
                return redirect('headmaster_fee_payment_create')

        context = {
            'form': form,
            'is_edit': pk is not None
        }
        return render(request, 'headmaster_template/add_fee_payment_form.html', context)
    

@method_decorator(login_required, name='dispatch')
class MadrasatulFeePaymentCreateUpdateView(View):
    def get(self, request, pk=None, *args, **kwargs):
        try:
            current_staff = Staffs.objects.get(admin=request.user)
            staff_branch = current_staff.branch
        except Staffs.DoesNotExist:
            return redirect('some_error_page')  # Replace with an appropriate error page

        # Get students from the branch of the current staff
        branch_students = Students.objects.filter(branch=staff_branch)

        if pk:
            madrasatul_fee_payment = get_object_or_404(MadrasatulFeePayment, pk=pk)
            form = MadrasatulFeePaymentForm(instance=madrasatul_fee_payment, branch_students=branch_students)
            context = {
                'form': form,
                'is_edit': True
            }
        else:
            form = MadrasatulFeePaymentForm(branch_students=branch_students)
            context = {
                'form': form,
                'is_edit': False
            }
        
        return render(request, 'headmaster_template/add_madrasatul_fee_payment_form.html', context)

    def post(self, request, pk=None, *args, **kwargs):
        try:
            current_staff = Staffs.objects.get(admin=request.user)
            staff_branch = current_staff.branch
        except Staffs.DoesNotExist:
            return redirect('some_error_page')  # Replace with an appropriate error page

        # Get students from the branch of the current staff
        branch_students = Students.objects.filter(branch=staff_branch)

        if pk:
            madrasatul_fee_payment = get_object_or_404(MadrasatulFeePayment, pk=pk)
            form = MadrasatulFeePaymentForm(request.POST, instance=madrasatul_fee_payment, branch_students=branch_students)
        else:
            form = MadrasatulFeePaymentForm(request.POST, branch_students=branch_students)

        if form.is_valid():
            fee_payment = form.save()
            action = request.POST.get('action')
            if action == 'save':
                messages.success(request, 'Madrasatul Fee Payment has been saved successfully!')
                return redirect('headmaster_madrasatul_fee_payment_list')
            elif action == 'save_and_continue':
                messages.success(request, 'Madrasatul Fee Payment has been saved! You can add another one.')
                return redirect('headmaster_madrasatul_fee_payment_create')

        context = {
            'form': form,
            'is_edit': pk is not None
        }
        return render(request, 'headmaster_template/add_madrasatul_fee_payment_form.html', context)
    
    

class MadrasatulFeePaymentListView(ListView):
    model = MadrasatulFeePayment
    template_name = 'headmaster_template/madrasatul_fee_payment_list.html'
    context_object_name = 'fee_payments'
    paginate_by = 10  # Optional: to paginate the list view

    def get_queryset(self):
        queryset = super().get_queryset()
        # You can add filters here if needed
        return queryset        
    
  
    
class MadrasatulFeePaymentDeleteView(View):
    def post(self, request, pk, *args, **kwargs):
        fee_payment = get_object_or_404(MadrasatulFeePayment, pk=pk)
        fee_payment.delete()
        return redirect('headmaster_madrasatul_fee_payment_list')               
    
@method_decorator(login_required, name='dispatch')    
class TransportFeePaymentCreateUpdateView(View):
    def get(self, request, pk=None, *args, **kwargs):
        # Retrieve the current staff member and their branch
        try:
            current_staff = Staffs.objects.get(admin=request.user)
            staff_branch = current_staff.branch
        except Staffs.DoesNotExist:
            return redirect('some_error_page')  # Replace with an appropriate error page

        # Get students from the branch of the current staff
        branch_students = Students.objects.filter(branch=staff_branch)

        if pk:
            transport_fee_payment = get_object_or_404(TransportFeePayment, pk=pk)
            form = TransportFeePaymentForm(instance=transport_fee_payment, branch_students=branch_students)
            context = {
                'form': form,
                'is_edit': True
            }
        else:
            form = TransportFeePaymentForm(branch_students=branch_students)
            context = {
                'form': form,
                'is_edit': False
            }
        
        return render(request, 'headmaster_template/add_transport_fee_payment_form.html', context)

    def post(self, request, pk=None, *args, **kwargs):
        # Retrieve the current staff member and their branch
        try:
            current_staff = Staffs.objects.get(admin=request.user)
            staff_branch = current_staff.branch
        except Staffs.DoesNotExist:
            return redirect('some_error_page')  # Replace with an appropriate error page

        # Get students from the branch of the current staff
        branch_students = Students.objects.filter(branch=staff_branch)

        if pk:
            transport_fee_payment = get_object_or_404(TransportFeePayment, pk=pk)
            form = TransportFeePaymentForm(request.POST, instance=transport_fee_payment, branch_students=branch_students)
        else:
            form = TransportFeePaymentForm(request.POST, branch_students=branch_students)

        if form.is_valid():
            transport_fee_payment = form.save()
            action = request.POST.get('action')
            if action == 'save':
                messages.success(request, 'Transport Fee Payment has been saved successfully!')
                return redirect('headmaster_transport_fee_payment_list')
            elif action == 'save_and_continue':
                messages.success(request, 'Transport Fee Payment has been saved! You can add another one.')
                return redirect('headmaster_transport_fee_payment_create')

        context = {
            'form': form,
            'is_edit': pk is not None
        }
        return render(request, 'headmaster_template/add_transport_fee_payment_form.html', context)
    
class TransportFeePaymentListView(ListView):
    model = TransportFeePayment
    template_name = 'headmaster_template/transport_fee_payment_list.html'
    context_object_name = 'transport_fee_payments'
    paginate_by = 10       
    
class TransportFeePaymentDeleteView(View):
    def post(self, request, pk, *args, **kwargs):
        fee_payment = get_object_or_404(TransportFeePayment, pk=pk)
        fee_payment.delete()
        return redirect('headmaster_transport_fee_payment_list') 
          
    
class ResultCreateUpdateView(LoginRequiredMixin, View):
    def get(self, request, exam_type_id, class_level_id, pk=None, *args, **kwargs):
        exam_type = get_object_or_404(ExamType, pk=exam_type_id)
        class_level = get_object_or_404(ClassLevel, pk=class_level_id)
        
        # Get the branch of the currently logged-in staff
        branch = request.user.staffs.branch if request.user.staffs else None

        # Filter subjects based on the class level
        subjects = Subject.objects.filter(class_level=class_level)

        if pk:
            result = get_object_or_404(Result, pk=pk)
            form = ResultForm(instance=result, exam_type=exam_type, class_level=class_level, branch=branch)
            context = {'form': form, 'is_edit': True}
        else:
            form = ResultForm(exam_type=exam_type, class_level=class_level, branch=branch)
            context = {'form': form, 'is_edit': False}
        
        context['subjects'] = subjects
        return render(request, 'headmaster_template/add_manage_result.html', context)

    def post(self, request, exam_type_id, class_level_id, pk=None, *args, **kwargs):
        exam_type = get_object_or_404(ExamType, pk=exam_type_id)
        class_level = get_object_or_404(ClassLevel, pk=class_level_id)

        # Get the branch of the currently logged-in staff
        branch = request.user.staffs.branch if request.user.staffs else None

        if pk:
            result = get_object_or_404(Result, pk=pk)
            form = ResultForm(request.POST, instance=result, exam_type=exam_type, class_level=class_level, branch=branch)
        else:
            form = ResultForm(request.POST, exam_type=exam_type, class_level=class_level, branch=branch)

        if form.is_valid():
            result = form.save(commit=False)
            result.exam_type = exam_type
            result.class_level = class_level
            result.save()
            action = request.POST.get('action')
            if action == 'save':
                messages.success(request, 'Result saved successfully!')
                return redirect('headmaster_view_student_to_add_result', exam_type_id=exam_type_id, class_level_id=class_level_id)  # Replace with your URL name for the result list view
            elif action == 'save_and_continue':
                messages.success(request, 'Result saved! You can add another one.')
                return redirect('headmaster_add_result', exam_type_id=exam_type_id, class_level_id=class_level_id)

        context = {'form': form, 'is_edit': pk is not None}
        return render(request, 'headmaster_template/add_manage_result.html', context)
    

class StudentResultCreateUpdateView(CreateView, UpdateView):
    model = Result
    form_class = StudentResultForm
    template_name = 'headmaster_template/add_student_result.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        student = get_object_or_404(Students, id=self.kwargs['student_id'])
        exam_type = get_object_or_404(ExamType, id=self.kwargs['exam_type_id'])
        kwargs.update({
            'student': student,
            'exam_type': exam_type
        })
        return kwargs

    def form_valid(self, form):
        # Set the `student`, `exam_type`, and `class_level` fields
        form.instance.student = get_object_or_404(Students, id=self.kwargs['student_id'])
        form.instance.exam_type = get_object_or_404(ExamType, id=self.kwargs['exam_type_id'])
        form.instance.class_level = form.instance.student.class_level  # Set class_level from student

        return super().form_valid(form)

    def get_object(self, queryset=None):
        # Return the existing Result object if in update mode
        if self.kwargs.get('pk'):
            return super().get_object(queryset)
        return None

    def get_success_url(self):
        action = self.request.POST.get('action')
        
        if action == 'save_and_continue':
            return reverse('headmaster_student_result_create_update', kwargs={
                'student_id': self.kwargs['student_id'],
                'exam_type_id': self.kwargs['exam_type_id'],
                'class_level_id': self.kwargs['class_level_id']
            })
        
        elif action == 'save' :
             return reverse('headmaster_view_student_to_add_result', kwargs={
            'exam_type_id': self.kwargs['exam_type_id'],
            'class_level_id': self.kwargs['class_level_id']
        })   
       

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['student'] = get_object_or_404(Students, id=self.kwargs['student_id'])
        context['exam_type'] = get_object_or_404(ExamType, id=self.kwargs['exam_type_id'])
        context['is_edit'] = self.object is not None
        return context
    
    
def get_subject_count_per_class_level(request):
    data = Subject.objects.values('class_level__class_name').annotate(subject_count=models.Count('id'))
    return JsonResponse(list(data), safe=False)    